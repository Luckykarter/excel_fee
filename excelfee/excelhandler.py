import openpyxl
from pycel import ExcelCompiler
from excelfee.models import Output, Input, OutputFile
import os
import base64


class ExcelHandler:
    BASE_PATH = os.path.dirname(os.path.abspath(__file__))
    TEMP_FILE = os.path.join(BASE_PATH, "temp.xlsx")

    def __init__(self, file):
        self.book = openpyxl.load_workbook(file)
        self.excel = ExcelCompiler(excel=self.book)

    def _set_value(self, sheet: str, cell: str, value: any):
        sheet_idx = 0 if not sheet else self.book.sheetnames.index(sheet)
        self.book.worksheets[sheet_idx][cell] = value

    def _get_value(self, sheet: str, cell: str):
        return self.excel.evaluate(f'{sheet}!{cell}')

    def _get_file_content(self):
        self.book.save(ExcelHandler.TEMP_FILE)
        with open(ExcelHandler.TEMP_FILE, "rb") as excel_file:
            content = base64.b64encode(excel_file.read())
            content = content.decode('utf-8')
        os.remove(ExcelHandler.TEMP_FILE)

        return content

    def calculate(self, data: Input,
                  excel_id: str, output_type='cells'):
        result = None
        try:
            for cell in data.input:
                self._set_value(cell.sheet, cell.cell, cell.value)

            cells = []
            if output_type == 'cells':
                self.excel = ExcelCompiler(excel=self.book)
                result = Output()
                for cell in data.output:
                    cell.value = self._get_value(cell.sheet, cell.cell)
                    cells.append(cell)
                result.cells = cells
                result.filename = excel_id
            elif output_type == 'file':
                result = OutputFile()
                result.content = self._get_file_content()

        except Exception as e:
            raise e

        return result

    @property
    def get_cells(self):
        result = {}

        for i, sheet_name in enumerate(self.book.sheetnames):
            sheet = self.book.worksheets[i]
            names = True
            cells = []
            cols = [col for col in sheet.iter_cols()]
            for c, col in enumerate(cols):
                if not names:
                    names = True
                    continue

                for n, cell in enumerate(col):
                    if cell.value is None:
                        continue
                    names = False
                    if c < len(cols) - 1:
                        tar_cell = cols[c + 1][n]
                        cells.append({
                            'cell': cell.value,
                            'type': 'N' if tar_cell.data_type == 'n' else 'T',
                            'coordinate': tar_cell.coordinate,
                        })

            result[sheet_name] = cells

        return result
