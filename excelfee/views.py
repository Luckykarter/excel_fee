from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework import status
import excelfee.serializers as serializers
from excelfee.excelhandler import ExcelHandler
import excelfee.models as models
import base64
import os
from django.core.cache import cache
from openpyxl import Workbook
import json


def _err(e=None):
    if e is None:
        e = "description of the error"
    return {'error': str(e)}


def _get_examples(e):
    return {
        "application/json": e
    }


def _get_html_list(items_list):
    res = '<ul>\n'
    for item in items_list:
        res += '<li>' + str(item) + '</li>\n'
    res += '</ul>'

    return res


def _calculate(request, calc_type):
    if not request.user.is_authenticated:
        return Response(status=status.HTTP_403_FORBIDDEN)

    req_data = request.stream.read()
    req_data = req_data.decode('windows-1252')
    req = json.loads(req_data)

    file_id = req.get('id')
    excel = cache.get(file_id)

    if excel is None:
        return Response(_err(f'File with id {file_id} not found'),
                        status=status.HTTP_404_NOT_FOUND)

    try:
        input_ = req.pop('input')
        input_data = models.Input()
        input_data.input = [models.Cell(**x) for x in input_.get('cells')]

        if calc_type == 'cells':
            output_ = req.pop('output')
            input_data.output = [models.Cell(**x) for x in output_.get('cells')]

        result = excel.calculate(input_data, file_id, calc_type)
        if calc_type == "cells":
            return Response(serializers.OutputSerializer(result).data,
                            status=status.HTTP_200_OK)
        elif calc_type == "file":
            return Response(serializers.OutputFileSerializer(result).data,
                            status=status.HTTP_200_OK)
        else:
            return Response(_err("Unknown calc_type: " + calc_type),
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    except FileNotFoundError as e:
        return Response({'error': str(e)},
                        status=status.HTTP_404_NOT_FOUND)

    except Exception as e:
        return Response({'error': str(e)},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    methods=['POST'],
    operation_id='calculate cells',
    operation_description='Calculate cells using MS Excel file\n'
                          'Update Excel cells with the values from JSON (array cells) '
                          'and get result as a calculated cells. Updated Excel file handled in-memory\n'
                          'To update/upload new Excel - use endpoint <b>upload_excel</b>\n\n'
                          '<ul>\n'
                          '<li>Input JSON object contains cells with values that should be populated</li>\n'
                          '<li>Output JSON object contains pointers to the cells that have to be returned as'
                          'calculation answer</li>'
                          '</ul>',
    request_body=serializers.InputSerializer,
    responses={200: openapi.Response('OK', serializers.OutputSerializer),
               404: openapi.Response('Excel file not found', serializers.ErrorSerializer,
                                     examples=_get_examples(
                                         _err("[Errno 2] No such file or directory: FeeCalcDemo.xlsx"))),
               500: openapi.Response('Calculation failed', serializers.ErrorSerializer,
                                     examples=_get_examples(_err()))})
@api_view(['POST'])
def calculate_cells(request):
    return _calculate(request, "cells")


@swagger_auto_schema(
    methods=['POST'],
    operation_id='calculate file',
    operation_description='Calculate cells using MS Excel file\n'
                          'Update Excel cells with the values from JSON (array cells) '
                          'and get result as a Base64 encoded file.'
                          'To update/upload new Excel - use endpoint <b>upload_excel</b>\n\n'
                          '<ul>\n'
                          '<li>Input JSON object contains cells with values that should be populated</li>\n'
                          '<li>Output JSON object contains Base64 encoded Excel file</li>'
                          '</ul>',
    request_body=serializers.InputFileSerializer,
    responses={200: openapi.Response('OK', serializers.OutputFileSerializer),
               404: openapi.Response('Excel file not found', serializers.ErrorSerializer,
                                     examples=_get_examples(
                                         _err("[Errno 2] No such file or directory: FeeCalcDemo.xlsx"))),
               500: openapi.Response('Calculation failed', serializers.ErrorSerializer,
                                     examples=_get_examples(_err()))})
@api_view(['POST'])
def calculate_file(request):
    return _calculate(request, "file")


@swagger_auto_schema(
    methods=['POST'],
    operation_id='upload_excel',
    operation_description='Upload MS Excel file',
    request_body=serializers.LoadExcelSerializer,
    responses={200: openapi.Response('File Uploaded Successfully', serializers.ExcelSerializer),
               400: openapi.Response('Bad Request', serializers.ErrorSerializer,
                                     examples=_get_examples(_err('content missing'))),
               500: openapi.Response('Error uploading Excel', serializers.ErrorSerializer,
                                     examples=_get_examples(_err('description of the error')))})
@api_view(['POST'])
def upload_excel(request):
    if not request.user.is_authenticated:
        return Response(status=status.HTTP_403_FORBIDDEN)

    try:
        content = request.data.get('content')
        file_id = request.data.get('id')
        folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
        if not os.path.exists(folder):
            os.mkdir(folder)
        filename = os.path.join(folder, 'temp.xlsx')

        if content:
            file_content = base64.b64decode(content)
            with open(filename, "wb") as f:
                f.write(file_content)
        else:
            new_book = Workbook()
            new_book.save(filename)

        excel = ExcelHandler(filename)
        cache.set(file_id, excel, 1800)
        return Response(status=status.HTTP_200_OK)

    except Exception as e:
        return Response(_err(e), status=status.HTTP_500_INTERNAL_SERVER_ERROR)


SUPPORTED_PROPERTIES = {
    'sheetnames', 'cells'
}


@swagger_auto_schema(
    methods=['GET'],
    operation_id='get_file_property',
    operation_description='Returns given property of the given Excel filename.\nSupported properties:\n'
                          '{}'.format(_get_html_list(SUPPORTED_PROPERTIES)),
    responses={200: openapi.Response('OK', serializers.PropertySerializer,
                                     examples=_get_examples({'property_name': ['value']})),
               400: openapi.Response('Bad Request', serializers.ErrorSerializer, examples=_get_examples(_err())),
               404: openapi.Response('Excel file not found', serializers.ErrorSerializer,
                                     examples=_get_examples(
                                         _err("[Errno 2] No such file or directory: FeeCalcDemo.xlsx"))),
               })
@api_view(['GET'])
def get_property(request, **kwargs):
    if not request.user.is_authenticated:
        return Response(status=status.HTTP_403_FORBIDDEN)

    file_id = kwargs.get('id')
    property = kwargs.get('property')
    property = property.lower()
    if not file_id:
        return Response(_err("file_id is empty"), status=status.HTTP_400_BAD_REQUEST)

    if not property:
        return Response(_err("property is empty"), status=status.HTTP_400_BAD_REQUEST)

    excel = cache.get(file_id)
    if excel is None:
        return Response(_err(f"file {file_id} not found"), status=status.HTTP_404_NOT_FOUND)

    if property not in SUPPORTED_PROPERTIES:
        resp = _err(f"property {property} is not supported")
        resp['supported_properties'] = [x for x in SUPPORTED_PROPERTIES]

        return Response(resp, status=status.HTTP_400_BAD_REQUEST)

    if property == 'sheetnames':
        return Response({'sheetnames': excel.book.sheetnames}, status=status.HTTP_200_OK)
    elif property == 'cells':
        return Response(excel.get_cells, status=status.HTTP_200_OK)
